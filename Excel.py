import openpyxl as opx
import os
import re

class Execl():

    def calculator(self,starttime, stoptime):
        start_h = int(starttime[:2])
        start_m = int(starttime[3:])
        stop_h = int(stoptime[:2])
        stop_m = int(stoptime[3:])

        try:
            self.relaxtime = int(input("请输入休息时间,默认休息时间为0"))
        except:
            self.relaxtime = 0

        h_sub = stop_h - start_h
        m_sub = stop_m - start_m
        if m_sub < 0:
            h_sub = h_sub - 1
        if h_sub < 0:
            night_y = True
            h_sub = 24 + h_sub
            while True:
                try:
                    night = input("检测到您存在夜班，是否工资不同？如果不同，请输入加倍后的工资")
                except:
                    yes_or_no = input("您确定真的没有？[Y/N]")
                    if yes_or_no == 'Y':
                        continue
                    else:
                        night_y = False
                        break
                else:
                    self.price = night

        try:
            trup = input("当天是否三薪[Y/N]")
        except:
            trup = 'N'

        if trup == 'Y':
            return [(h_sub + abs(m_sub) / 60 - self.relaxtime / 60) * 3,night_y]
        else:
            return [h_sub + abs(m_sub) / 60 - self.relaxtime / 60,night_y]

    def readme(self):
        print(f"工资计算器\n使用说明："
              f"\n1.把你的Excel文档放到这个文件夹下\n"
              f"2.运行程序，复制文件名到黑框里（将文件后缀名一起复制）\n"
              f"3.回车按照提示一步步操作即可")
        j = os.system("pause")
        i = os.system('cls')

    def find_data(self):
        self.readme()
        while True:
            try:
                file_name = input("请输入文件名")
                wb = opx.load_workbook(file_name)
            except:
                print("输入的文件名错误或者为空")
                continue
            else:
                break

        ws = wb.active

        col_b = ws['B']

        while True:
            try:
                input_value = input("请输入你想查找的名字！！")
            except:
                continue
            else:
                break

        i = 0

        for item in col_b:
            if item.value == input_value:
                break
            i = i + 1
        i = i + 1

        line_i = ws[i]
        self.line_v = list()

        for item in line_i:
            around = re.search(r'\d',item.value).span()
            value_s = item.value()
            value = value_s[around[0]:]
            self.line_v.append(value)

        del self.line_v[0]
        self.line_v.remove(input_value)

        for i in range(len(self.line_v)):
            try:
                self.line_v.remove('')
            except:
                break
        for i in range(len(self.line_v)):
            try:
                self.line_v.remove(None)
            except:
                break

        self.result = list()
        self.lenth = len(self.line_v)
        for i in range(0,self.lenth , 2):
            self.result.append(self.calculator(self.line_v[i], self.line_v[i + 1]))

        while True:
            try:
                self.price = int(input("请输入时薪"))
            except:
                continue
            else:
                break
        i = 1

        self.locale_price = list()
        self.night = list()
        for item in self.result:
            print(f'第{i}日工资为:{item[0] * self.price}')
            self.locale_price.append(item[0]*self.price)
            self.night.append(item[1])
            i += 1


        k = os.system("pause")




