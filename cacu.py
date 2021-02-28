import datetime
import openpyxl
import tkinter.messagebox
import tkinter

class cacu():

    def __init__(self):
        self.root = tkinter.Tk()
        self.root.withdraw()

    def get_start_time(self,start_hours, start_minuit):
        starttime = datetime.time(start_hours, start_minuit)
        self.start = datetime.time.strftime(starttime, "%H:%M")

    def get_stop_time(self,stop_hours, stop_minuit):
        stoptime = datetime.time(stop_hours, stop_minuit)
        self.stop = datetime.time.strftime(stoptime, "%H:%M")

    def get_relax_time(self,relax):
        self.relax = relax
        self.relax_time = relax / 60

    def getdelta(self,start, stop):
        hourdelta = int(stop[:2]) - int(start[:2])
        minuitdelta = int(stop[3:]) - int(start[3:])
        if minuitdelta < 0:
            hourdelta = hourdelta - 1

        if hourdelta < 0:
            hourdelta = 24 + hourdelta

        timedel = hourdelta + abs(minuitdelta) / 60
        return timedel

    def cacuresult(self, price,Triple,choose,t_o_p_value):
        self.timedel = self.getdelta(self.start,self.stop)
        self.result = price * (self.timedel - self.relax_time)

        self.night = False

        if t_o_p_value == '':
            pass
        else:
            t_o_p_value = float(t_o_p_value)

        if Triple == 1:
            self.result = self.result * 3
        if choose == 1:
            self.result = self.result * t_o_p_value
            self.night = True
        elif choose == 2:
            self.result = self.result + t_o_p_value * self.timedel
            self.night = True

    def writer_to_excel(self,start=None,stop=None,relax=None,result=None,night=False):

        try:
            wb = openpyxl.load_workbook('My Table.xlsx')
        except:
            tkinter.messagebox.showinfo('创建文件','Excel文件不存在！！\n点击确定创建文件')
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active

            self.ws.append(['日期', '开始时间', '结束时间','休息时间','是否夜班', '当日工资'])
            tkinter.messagebox.showinfo('创建文件', 'Excel文件创建成功')

        else:
            self.wb = openpyxl.load_workbook('My Table.xlsx')
            tkinter.messagebox.showinfo('文件存在','Excel文件存在，打开成功！！')

            self.ws = self.wb.active

        date = datetime.datetime.now()
        date = datetime.datetime.strftime(date, '%Y-%m-%d')

        if result is None:
            col = [date, self.start, self.stop, self.night ,self.relax,self.result]
            tkinter.messagebox.showinfo('工资结果', f'您当天的工资是{self.result}元')
        else:
            col = [date, start, stop, relax, night, result]


        self.ws.append(col)

        self.wb.save('My Table.xlsx')



